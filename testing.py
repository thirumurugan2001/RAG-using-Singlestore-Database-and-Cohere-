import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import json

def generate_rag_excel_file(data, filename="rag_data.xlsx"):
    """
    Generate an Excel file from RAG data with formatting
    
    Args:
        data: List of dictionaries containing RAG data
        filename: Output Excel filename (default: "rag_data.xlsx")
    
    Returns:
        str: Path to the generated Excel file
    """
    try:
        # Convert data to DataFrame
        df = pd.DataFrame(data)
        
        # Create Excel writer object
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='RAG Data', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['RAG Data']
            
            # Apply formatting
            apply_excel_formatting(worksheet, df)
        
        print(f"Excel file '{filename}' generated successfully!")
        print(f"Total records: {len(data)}")
        return filename
        
    except Exception as e:
        print(f"Error generating Excel file: {e}")
        return None

def apply_excel_formatting(worksheet, df):
    """
    Apply formatting to the Excel worksheet
    """
    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell_alignment = Alignment(wrap_text=True, vertical='top')
    
    # Format header row
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Format data cells
    for row in range(2, len(df) + 2):  # +2 because Excel is 1-indexed and we have header
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.alignment = cell_alignment
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        worksheet.column_dimensions[column_letter].width = adjusted_width

def generate_rag_excel_with_summary(data, filename="rag_data_with_summary.xlsx"):
    """
    Generate an Excel file with RAG data and a summary sheet
    """
    try:
        # Create workbook
        workbook = Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Create main data sheet
        ws_data = workbook.create_sheet("RAG Data")
        
        # Convert to DataFrame for easier handling
        df = pd.DataFrame(data)
        
        # Write headers
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            ws_data.cell(row=1, column=col_idx, value=header)
        
        # Write data
        for row_idx, row_data in enumerate(data, 2):
            ws_data.cell(row=row_idx, column=1, value=row_data['sno'])
            ws_data.cell(row=row_idx, column=2, value=row_data['description'])
        
        # Create summary sheet
        ws_summary = workbook.create_sheet("Summary")
        
        # Add summary information
        summary_data = [
            ["RAG Data Summary", ""],
            ["Total Records", len(data)],
            ["Data Categories", "Professional Profile, Skills, Projects, Experience"],
            ["Generated On", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["", ""],
            ["Key Categories Count", ""],
            ["AI/ML Skills", sum(1 for item in data if any(keyword in item['description'].lower() for keyword in ['ai', 'machine learning', 'llm', 'rag']))],
            ["Projects", sum(1 for item in data if 'project' in item['description'].lower())],
            ["Technologies", sum(1 for item in data if any(keyword in item['description'].lower() for keyword in ['python', 'javascript', 'react', 'aws', 'azure']))],
            ["Experience", sum(1 for item in data if any(keyword in item['description'].lower() for keyword in ['experience', 'worked', 'role', 'position']))]
        ]
        
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                ws_summary.cell(row=row_idx, column=col_idx, value=value)
        
        # Apply formatting
        apply_excel_formatting(ws_data, df)
        
        # Format summary sheet
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col in range(1, 3):
            cell = ws_summary.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
        
        # Auto-adjust column widths for summary
        for column in ws_summary.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 30)
            ws_summary.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        workbook.save(filename)
        print(f"Excel file '{filename}' with summary generated successfully!")
        print(f"Total records: {len(data)}")
        return filename
        
    except Exception as e:
        print(f"Error generating Excel file: {e}")
        return None

# Example usage
if __name__ == "__main__":
    # Your RAG data here (copy the ragData array from your question)
    
    # Generate basic Excel file
    generate_rag_excel_file(ragData, "thirumurugan_rag_data.xlsx")
    
    # Generate Excel file with summary
    generate_rag_excel_with_summary(ragData, "thirumurugan_rag_data_detailed.xlsx")
    
    # You can also generate CSV for quick viewing
    df = pd.DataFrame(ragData)
    df.to_csv("rag_data_backup.csv", index=False)
    print("CSV backup file 'rag_data_backup.csv' generated!")